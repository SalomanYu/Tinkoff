import json
import xlrd
import xlsxwriter
import openpyxl

vacancies_groups = json.load(open('Jsons/1FinalResult.json', 'r'))
requirements = json.load(open('Jsons/AMOUNT_requirements.json', 'r'))

def set_up_mathes(excel):
    writer_wb = openpyxl.load_workbook(excel)
    writer_sh = writer_wb.worksheets[0]
    
    workbook = xlrd.open_workbook(excel)
    sheet = workbook.sheet_by_index(0)
    for group in vacancies_groups:
        for item in range(1, sheet.nrows-1):
            if group['id'] == int(sheet.cell(item, 0).value):
                for req_group in  requirements:
                    if req_group['Группа'] == group['Группа']:
                        req = req_group['Требования']

                        for title_col in range(sheet.ncols):
                            if sheet.cell(0, title_col).value in req:
                                writer_sh.cell(item+1, title_col+1).value = 1
    writer_wb.save('Excel/requirements_excel.xlsx')


def create_frame_table():
    workbook = xlsxwriter.Workbook('Excel/requirements_excel.xlsx')
    worksheet = workbook.add_worksheet()

    worksheet.write(0, 0, 'Соответствия')
    worksheet.write(0, 1, 'Вес профессии в соответствии')
    worksheet.write(0, 2, 'Уровень должности')
    worksheet.write(0, 3, 'Вес профессии в уровне')
    worksheet.write(0, 4, 'Наименование професии и различные написания')

    col = 5
    for group in requirements:
        for req in group['Требования']:
            worksheet.write(0, col, req)
            col += 1

    row_count = 1
    for group in vacancies_groups:
        for vacance in group['Все вакансии']:
            try:
                weight_in_level = vacance['Вес в уровне']
            except KeyError:
                weight_in_level = 0
            if weight_in_level == 1:

                worksheet.write(row_count, 0, group['id'])
                worksheet.write(row_count, 1, vacance['Вес'])
                worksheet.write(row_count, 2, vacance['Уровень'])
                worksheet.write(row_count, 3, weight_in_level)
                worksheet.write(row_count, 4, vacance['Вакансия'])
                row_count += 1

    # set_up_mathes('Excel/requirements_excel.xlsx')
    workbook.close()

create_frame_table()
set_up_mathes('Excel/requirements_excel.xlsx')

def add_zeros(excel):
    workbook = xlrd.open_workbook(excel)
    sheet = workbook.sheet_by_index(0)

    writer_workbook = openpyxl.load_workbook(excel)
    writer_sheet = writer_workbook.worksheets[0]

    for num_col in range(5, sheet.ncols-5):
        for num_row in range(sheet.nrows):
            if writer_sheet.cell(num_row+1, num_col+1).value == None:
                writer_sheet.cell(num_row+1, num_col+1).value = 0
                print('добавили ноль')
            else:
                print(writer_sheet.cell(num_row+1, num_col+1).value)
    
    writer_workbook.save(excel)

add_zeros('Excel/requirements_excel.xlsx')